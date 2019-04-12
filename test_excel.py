import openpyxl

'''
# Excel file with list1:
name	id	result
paul	1	1
sveta	5	5
dan	    3	3
liza	4	4
'''

wb = openpyxl.load_workbook(filename = './excel_file.xlsx')

# active list
ws = wb.active
# or list name
ws = wb['list1']
# or list number
ws = wb.worksheets[0]

# rename list
#ws.title = "sheet1"

#считываем значение определенной ячейки
val = ws['A1'].value
print(val)


first_column = ws['B']
# Print the contents
for x in range(len(first_column)):
    if x!=0:
        print(str(x)+str(first_column[x].value))
        ws.cell(row=x+1, column=3).value  = first_column[x].value*2
    else:
        pass

# writing values to cells:
#e3
ws['E3']                        = 10
#e4
ws.cell(row=4, column=5).value  = 20
#e5
e5                              = ws.cell(row=5, column=5)
e5.value                        = 30

wb.save("excel_file.xlsx")